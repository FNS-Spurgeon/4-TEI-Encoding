from openpyxl import load_workbook
import xml.etree.ElementTree as eT
import os

xmlFiles = "../tei-files/vol2-1851-1875-xml"

wb = load_workbook(filename='../../3-DataExtraction/hathitrustLinks.xlsx')
pages = wb['Part3']

list_pages = []
for row in pages.iter_rows(min_row=2, max_row=pages.max_row, values_only=True):
    list_pages.append(row)

for file in os.listdir(xmlFiles):
    if file.endswith(".xml"):
        xml_path = os.path.join(xmlFiles, file)
        ns = {'tei': 'http://www.tei-c.org/ns/1.0'}
        eT.register_namespace('', 'http://www.tei-c.org/ns/1.0')

        tree = eT.parse(xml_path)
        root = tree.getroot()

        for i in range(len(list_pages)):
            for biblScope in root.findall(".//tei:biblScope[@unit='page']", ns):
                if "-" in biblScope.text:
                    pageBeginning = biblScope.text.split("-")[0]
                    if pageBeginning == str(list_pages[i][0]):
                        biblScope.set("source", list_pages[i][1])
                    else:
                        pass
                else:
                    if biblScope.text == str(list_pages[i][0]):
                        biblScope.set("source", list_pages[i][1])

        tree.write(xml_path, encoding="UTF-8", xml_declaration=True)

        print(file, ": DONE")
